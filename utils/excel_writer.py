"""Professional Excel workbook export."""
from __future__ import annotations

from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils.helpers import STANDARD_COLUMNS, ensure_schema, sanitize_sheet_name

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    df = ensure_schema(df)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([None if pd.isna(value) else value for value in row])

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN_GRAY)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    if ws.max_row > 1:
        table_name = "Tbl_" + "".join(ch for ch in ws.title if ch.isalnum())[:20]
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)

    header_index = {cell.value: cell.column for cell in ws[1]}
    for col_name in ("Price", "Change"):
        col = header_index.get(col_name)
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '$#,##0.00;[Red]-$#,##0.00'
    for col_name in ("Volume", "Market Cap"):
        col = header_index.get(col_name)
        if col:
            for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in cell:
                    c.number_format = '#,##0'
    change_col = header_index.get("Change %")
    if change_col and ws.max_row > 1:
        letter = get_column_letter(change_col)
        for cell in ws[letter][1:]:
            cell.number_format = '0.00"%";[Red]-0.00"%"'
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
    ws.row_dimensions[1].height = 24


def build_summary(frames: Mapping[str, pd.DataFrame]) -> pd.DataFrame:
    non_empty = [ensure_schema(df) for df in frames.values() if df is not None and not df.empty]
    if not non_empty:
        return pd.DataFrame(columns=STANDARD_COLUMNS)
    summary = pd.concat(non_empty, ignore_index=True)
    summary["Change %"] = pd.to_numeric(summary["Change %"], errors="coerce")
    summary = summary.drop_duplicates(subset=["Source", "Category", "Symbol"], keep="first")
    return summary.sort_values("Change %", ascending=False, na_position="last").reset_index(drop=True)


def export_workbook(frames: Mapping[str, pd.DataFrame], output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    ordered_sheets = [
        "Robinhood_Gainers", "Robinhood_Losers", "Robinhood_Actives",
        "FMP_Gainers", "FMP_Losers", "FMP_Actives",
        "Polygon_Gainers", "Polygon_Losers",
        "Yahoo_Market", "Finviz_Market",
    ]
    for sheet_name in ordered_sheets:
        ws = wb.create_sheet(sanitize_sheet_name(sheet_name))
        _write_dataframe(ws, frames.get(sheet_name, pd.DataFrame(columns=STANDARD_COLUMNS)))

    summary_ws = wb.create_sheet("Summary")
    _write_dataframe(summary_ws, build_summary(frames))
    summary_ws.sheet_properties.tabColor = "70AD47"
    summary_ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
